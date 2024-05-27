from psychopy import visual, core, event
from openpyxl import*
from openpyxl.styles import Font
import time


screen_width = 1920
screen_height = 1080

win_width = screen_width // 2
win_height = screen_height
win_pos = (0, 0)
win = visual.Window(size=(screen_width, screen_height), screen=1, fullscr=False, allowGUI=True, units='pix',
                    color='black', pos=win_pos)

def time_ms(startTime):
    currentTime_ns = time.time_ns()
    print('currentTime_ns', currentTime_ns)
    print('startTime',startTime)
    wantedTime_ns = currentTime_ns - startTime
    wantedTime_ms = wantedTime_ns/1e6
    print('wantedTime_ms',wantedTime_ms)
    return(wantedTime_ms)

workbook = load_workbook(filename='hexaco_P1.xlsx')
sheet = workbook.worksheets[0]
alphabet_with_ones = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1', 'N1', 'O1', 'P1', 'Q1', 'R1', 'S1', 'T1', 'U1', 'V1', 'W1', 'X1', 'Y1', 'Z1']
alphabet_with_twos = ['A2', 'B2', 'C2', 'D2', 'E2', 'F2', 'G2', 'H2', 'I2', 'J2', 'K2', 'L2', 'M2', 'N2', 'O2', 'P2', 'Q2', 'R2', 'S2', 'T2', 'U2', 'V2', 'W2', 'X2', 'Y2', 'Z2']
columns = []
current_participant = []
reverse_questions = [2,3,4,6,10,13,15,18,19,20]
text_responses = []

def reverse_scales():
    for cell in sheet:
        possible_columns = list(range(2,69))
        if cell.row in reverse_questions and cell.column in possible_columns:
            cell_to_reverse = cell.coordinate
            cell_row = cell.row
            cell_column = cell.column
            if int(cell_to_reverse.value) == 1:
                sheet.cell(row=cell_row, column=cell_column, value=5)
            if int(cell_to_reverse.value) == 2:
                sheet.cell(row=cell_row, column=cell_column, value=2)
            if int(cell_to_reverse.value) == 4:
                sheet.cell(row=cell_row, column=cell_column, value=2)
            if int(cell_to_reverse.value) == 5:
                sheet.cell(row=cell_row, column=cell_column, value=1)
        else:
            pass

def participant_number():
    index = 0
    font = Font(bold=True)
    for cell in sheet[1]:
        #print(cell)
        #print(cell.coordinate)
        if cell.coordinate in alphabet_with_ones:
            #print(cell.coordinate)
            cell_index = alphabet_with_ones.index(cell.coordinate)
            cell_column = sheet[alphabet_with_ones[cell_index]].column
            #print(cell_column)
            #print(type(cell_column))
            current_participant.append(cell_column)
            participant_now = sheet.cell(row=1, column=cell_column).value
            next_participant = sheet.cell(row=1, column=cell_column+1).value
            #print(participant_now)
            if cell_index == 0:
                sheet.cell(row=1, column=cell_column, value="Questions")
                sheet.cell(row=1, column=cell_column).font = font
            if cell_column == 1:
                if participant_now == 'Questions':
                    sheet.cell(row=1, column=cell_column+1, value="Participant1")
                    sheet.cell(row=1, column=cell_column+1).font = font
            if next_participant is None:
                index+=cell_column
                #print(sheet.cell(row=1, column=cell_column+1).value)
                sheet.cell(row=1, column=cell_column+1, value=f"Participant{index}")
                sheet.cell(row=1,column=cell_column+1).font = font
                break
        if cell.coordinate in alphabet_with_twos:
            # print(cell.coordinate)
            cell_index = alphabet_with_twos.index(cell.coordinate)
            cell_column = sheet[alphabet_with_twos[cell_index]].column
            # print(cell_column)
            # print(type(cell_column))
            current_participant.append(cell_column)
            participant_now = sheet.cell(row=1, column=cell_column).value
            next_participant = sheet.cell(row=1, column=cell_column + 1).value
            # print(participant_now)
            if cell_column == 1:
                if participant_now == 'Questions':
                    sheet.cell(row=1, column=cell_column + 1, value="Participant1")
                    sheet.cell(row=1, column=cell_column + 1).font = font
            if next_participant is None:
                index += cell_column
                # print(sheet.cell(row=1, column=cell_column+1).value)
                sheet.cell(row=1, column=cell_column+1, value=f"Participant{index}")
                sheet.cell(row=1, column=cell_column+1).font = font
                break

def get_empty_col():
    for cell in sheet[2]:
        if cell.coordinate in alphabet_with_twos:
            cell_index = alphabet_with_twos.index(cell.coordinate)
            cell_column = sheet[alphabet_with_twos[cell_index]].column
            column_now = sheet.cell(row=1, column=cell_column).value
            next_column = cell_column+1
            next_column_value=sheet.cell(row=2,column=next_column).value
            if column_now is not None and next_column_value is None:
                columns.append(next_column)
                break
        if cell.coordinate in alphabet_with_twos:
            cell_index = alphabet_with_twos.index(cell.coordinate)
            cell_column = sheet[alphabet_with_twos[cell_index]].column
            next_column = cell_column+1
            columns.append(next_column)



#Define questions and options
def welcome():
    sheet.cell(row=1, column=1)
    back_slide = visual.TextStim(win, text='Back', height=20, pos=[-375, -400])
    next_slide = visual.TextStim(win, text='Next', height=20, pos=[375, -400])
    back_button = visual.TextStim(win, text='(press LEFT)', height=20, pos=[-375,-425])
    next_button = visual.TextStim(win, text='(press RIGHT)', height=20, pos=[375,-425])
    not_at_all = visual.TextStim(win, text='(not at all)', height=20, pos=[-220,-275])
    very_much = visual.TextStim(win, text='(very much)', height=20, pos=[220,-275])
    options = [visual.TextStim(win, text='1', height=35, pos=[-220,-250]),
               visual.TextStim(win, text='2', height=35, pos=[-110, -250]),
               visual.TextStim(win, text="3", height=35, pos=[0, -250]),
               visual.TextStim(win, text='4', height=35, pos=[110, -250]),
               visual.TextStim(win, text='5', height=35, pos=[220, -250])]
    option1 =[visual.TextStim(win, text='1', height=35, pos=[-220,-250], color=[0, 255, 0]),
               visual.TextStim(win, text='2', height=35, pos=[-110, -250]),
               visual.TextStim(win, text="3", height=35, pos=[0, -250]),
               visual.TextStim(win, text='4', height=35, pos=[110, -250]),
               visual.TextStim(win, text='5', height=35, pos=[220, -250])]
    option2 =[visual.TextStim(win, text='1', height=35, pos=[-220,-250]),
               visual.TextStim(win, text='2', height=35, pos=[-110, -250], color=[0, 255, 0]),
               visual.TextStim(win, text="3", height=35, pos=[0, -250]),
               visual.TextStim(win, text='4', height=35, pos=[110, -250]),
               visual.TextStim(win, text='5', height=35, pos=[220, -250])]
    option3 =[visual.TextStim(win, text='1', height=35, pos=[-220,-250]),
               visual.TextStim(win, text='2', height=35, pos=[-110, -250]),
               visual.TextStim(win, text="3", height=35, pos=[0, -250], color=[0, 255, 0]),
               visual.TextStim(win, text='4', height=35, pos=[110, -250]),
               visual.TextStim(win, text='5', height=35, pos=[220, -250])]
    option4 =[visual.TextStim(win, text='1', height=35, pos=[-220,-250]),
               visual.TextStim(win, text='2', height=35, pos=[-110, -250]),
               visual.TextStim(win, text="3", height=35, pos=[0, -250]),
               visual.TextStim(win, text='4', height=35, pos=[110, -250], color=[0, 255, 0]),
               visual.TextStim(win, text='5', height=35, pos=[220, -250])]
    option5 =[visual.TextStim(win, text='1', height=35, pos=[-220,-250]),
               visual.TextStim(win, text='2', height=35, pos=[-110, -250]),
               visual.TextStim(win, text="3", height=35, pos=[0, -250]),
               visual.TextStim(win, text='4', height=35, pos=[110, -250]),
               visual.TextStim(win, text='5', height=35, pos=[220, -250], color=[0, 255, 0])]
    options_colored = [option1, option2, option3, option4, option5]
    welcome_stimuli = [visual.TextStim(win, text="You will now be asked to rate a couple of questions regarding some characteristics about yourself.\n\n You may go back and revise your answers. \n\n Press SPACE to start.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text="1. I rarely hold a grudge, even against people who have badly wronged me.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text="2. People sometimes tell me that I am too critical of others.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text="3. If I knew that I could never get caught, I would be willing to steal a million dollars.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text="4. People sometimes tell me that I'm too stubborn.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text="5. Having a lot of money is not especially important to me.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text="6. People think of me as someone who has a quick temper.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text="7. My attitude toward people who have treated me badly is 'forgive and forget'.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text="8. I wouldn't use flattery to get a raise or promotion at work, even if I thought it would succeed.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text="9. I am usually quite flexible in my opinions when people disagree with me.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text="10. If I want something from someone, I will laugh at that person's worst jokes.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text="11. For this question, please press 4.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text="12. I tend to be lenient in judging other people.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text ="13. I would never accept a bribe, even if it were very large.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text ="14. When people tell me that I’m wrong, my first reaction is to argue with them.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text ="15. Most people tend to get angry more quickly than I do.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text ="16. I think that I am entitled to more respect than the average person is.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text ="17. I wouldn't pretend to like someone just to get that person to do favors for me.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text ="18. Even when people make a lot of mistakes, I rarely say anything negative.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text ="19. I'd be tempted to use counterfeit money, if I were sure I could get away with it.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text ="20. I would get a lot of pleasure from owning expensive luxury goods.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text ="21. I want people to know that I am an important person of high status.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text ="This is the last slide, you may still go back and check your answers.\n\n If you are done, please PRESS SPACE to submit your answers.", height=30, pos=[0, 0], wrapWidth=500)
]
    stimuli_order = list(range(len(welcome_stimuli)))
    scale_order = [0,1,2,3,4]
    timesStimuli = []
    i = 0
    row_index = 2
    col_index = -1
    startTime = time.time_ns()
    timesStimuli.append(startTime)
    print('startTime', startTime)
    while i < len(stimuli_order):
        welcome_stimuli[stimuli_order[i]].draw()
        if i > 1:
            back_slide.draw()
            back_button.draw()
        if 0< i < len(stimuli_order)-1:
            not_at_all.draw()
            very_much.draw()
            [i.draw() for i in options]
            #next_slide.draw()
            #next_button.draw()
        win.flip()
        keys = event.getKeys(keyList=['left', 'space', '1', '2', '3', '4', '5'])
        if 'left' in keys and i > 1:
            i -= 1
            row_index -= 1
        if 'space' in keys and i == len(stimuli_order)-1:
            i += 1
        if 'space' in keys and i == 0:
            i += 1
            welcome_stimuli[stimuli_order[i]].draw()
            win.flip()
        if any(key in keys for key in ['1', '2', '3',  '4', '5']) and 0 < i < len(stimuli_order)-1:
            #print(welcome_stimuli[stimuli_order[i]])
            #print(keys[0])
            #print(columns)
            timePress_ms = time_ms(startTime)
            keyTime = str(str(keys[0]) + ','+str(timePress_ms))
            sheet.cell(row=row_index, column=1,value=row_index-1)
            sheet.cell(row=row_index, column=columns[col_index], value=keyTime)
            row_index += 1
            if [int(keys[0]) in scale_order]:
                scale_num = int(keys[0])-1
                if int(keys[0]) > 1:
                    [b.draw() for b in options_colored[scale_order[scale_num]]]
                if int(keys[0]) == 1:
                    [b.draw() for b in options_colored[scale_order[0]]]
                welcome_stimuli[stimuli_order[i]].draw()
                #[i.draw() for i in options]
                not_at_all.draw()
                very_much.draw()
                if i > 1:
                    back_slide.draw()
                    back_button.draw()
                #if i != len(stimuli_order) - 1:
                    #next_slide.draw()
                    #next_button.draw()
                win.flip()
                core.wait(1.0)
                i += 1

def bye():
    sheet.cell(row=1, column=1)
    not_at_all = visual.TextStim(win, text='(not at all)', height=20, pos=[-300,-250])
    extremely = visual.TextStim(win, text='  (extremely)', height=20, pos=[300,-250])
    options_positions = [(-225, -250), (-175, -250), (-125, -250), (-75, -250), (-25, -250),
                         (25, -250), (75, -250), (125, -250), (175, -250), (225, -250)]
    options = [visual.TextStim(win, text='1', height=35, pos=options_positions[0]),
               visual.TextStim(win, text="2", height=35, pos=options_positions[1]),
               visual.TextStim(win, text='3', height=35, pos=options_positions[2]),
               visual.TextStim(win, text='4', height=35, pos=options_positions[3]),
               visual.TextStim(win, text='5', height=35, pos=options_positions[4]),
               visual.TextStim(win, text='6', height=35, pos=options_positions[5]),
               visual.TextStim(win, text='7', height=35, pos=options_positions[6]),
               visual.TextStim(win, text='8', height=35, pos=options_positions[7]),
               visual.TextStim(win, text='9', height=35, pos=options_positions[8]),
               visual.TextStim(win, text='10', height=35, pos=options_positions[9])]

    option0 = [visual.TextStim(win, text='1', height=35, pos=options_positions[0]),
               visual.TextStim(win, text="2", height=35, pos=options_positions[1]),
               visual.TextStim(win, text='3', height=35, pos=options_positions[2]),
               visual.TextStim(win, text='4', height=35, pos=options_positions[3]),
               visual.TextStim(win, text='5', height=35, pos=options_positions[4]),
               visual.TextStim(win, text='6', height=35, pos=options_positions[5]),
               visual.TextStim(win, text='7', height=35, pos=options_positions[6]),
               visual.TextStim(win, text='8', height=35, pos=options_positions[7]),
               visual.TextStim(win, text='9', height=35, pos=options_positions[8]),
               visual.TextStim(win, text='10', height=35, pos=options_positions[9], color=[0, 255, 0])]

    option1 = [visual.TextStim(win, text='1', height=35, pos=options_positions[0], color=[0,255,0]),
               visual.TextStim(win, text="2", height=35, pos=options_positions[1]),
               visual.TextStim(win, text='3', height=35, pos=options_positions[2]),
               visual.TextStim(win, text='4', height=35, pos=options_positions[3]),
               visual.TextStim(win, text='5', height=35, pos=options_positions[4]),
               visual.TextStim(win, text='6', height=35, pos=options_positions[5]),
               visual.TextStim(win, text='7', height=35, pos=options_positions[6]),
               visual.TextStim(win, text='8', height=35, pos=options_positions[7]),
               visual.TextStim(win, text='9', height=35, pos=options_positions[8]),
               visual.TextStim(win, text='10', height=35, pos=options_positions[9])]

    option2 = [visual.TextStim(win, text='1', height=35, pos=options_positions[0]),
               visual.TextStim(win, text="2", height=35, pos=options_positions[1], color=[0,255,0]),
               visual.TextStim(win, text='3', height=35, pos=options_positions[2]),
               visual.TextStim(win, text='4', height=35, pos=options_positions[3]),
               visual.TextStim(win, text='5', height=35, pos=options_positions[4]),
               visual.TextStim(win, text='6', height=35, pos=options_positions[5]),
               visual.TextStim(win, text='7', height=35, pos=options_positions[6]),
               visual.TextStim(win, text='8', height=35, pos=options_positions[7]),
               visual.TextStim(win, text='9', height=35, pos=options_positions[8]),
               visual.TextStim(win, text='10', height=35, pos=options_positions[9])]

    option3 = [visual.TextStim(win, text='1', height=35, pos=options_positions[0]),
               visual.TextStim(win, text="2", height=35, pos=options_positions[1]),
               visual.TextStim(win, text='3', height=35, pos=options_positions[2], color=[0,255,0]),
               visual.TextStim(win, text='4', height=35, pos=options_positions[3]),
               visual.TextStim(win, text='5', height=35, pos=options_positions[4]),
               visual.TextStim(win, text='6', height=35, pos=options_positions[5]),
               visual.TextStim(win, text='7', height=35, pos=options_positions[6]),
               visual.TextStim(win, text='8', height=35, pos=options_positions[7]),
               visual.TextStim(win, text='9', height=35, pos=options_positions[8]),
               visual.TextStim(win, text='10', height=35, pos=options_positions[9])]

    option4 = [visual.TextStim(win, text='1', height=35, pos=options_positions[0]),
               visual.TextStim(win, text="2", height=35, pos=options_positions[1]),
               visual.TextStim(win, text='3', height=35, pos=options_positions[2]),
               visual.TextStim(win, text='4', height=35, pos=options_positions[3], color=[0, 255, 0]),
               visual.TextStim(win, text='5', height=35, pos=options_positions[4]),
               visual.TextStim(win, text='6', height=35, pos=options_positions[5]),
               visual.TextStim(win, text='7', height=35, pos=options_positions[6]),
               visual.TextStim(win, text='8', height=35, pos=options_positions[7]),
               visual.TextStim(win, text='9', height=35, pos=options_positions[8]),
               visual.TextStim(win, text='10', height=35, pos=options_positions[9])]

    option5 = [visual.TextStim(win, text='1', height=35, pos=options_positions[0]),
               visual.TextStim(win, text="2", height=35, pos=options_positions[1]),
               visual.TextStim(win, text='3', height=35, pos=options_positions[2]),
               visual.TextStim(win, text='4', height=35, pos=options_positions[3]),
               visual.TextStim(win, text='5', height=35, pos=options_positions[4], color=[0, 255, 0]),
               visual.TextStim(win, text='6', height=35, pos=options_positions[5]),
               visual.TextStim(win, text='7', height=35, pos=options_positions[6]),
               visual.TextStim(win, text='8', height=35, pos=options_positions[7]),
               visual.TextStim(win, text='9', height=35, pos=options_positions[8]),
               visual.TextStim(win, text='10', height=35, pos=options_positions[9])]

    option6 = [visual.TextStim(win, text='1', height=35, pos=options_positions[0]),
               visual.TextStim(win, text="2", height=35, pos=options_positions[1]),
               visual.TextStim(win, text='3', height=35, pos=options_positions[2]),
               visual.TextStim(win, text='4', height=35, pos=options_positions[3]),
               visual.TextStim(win, text='5', height=35, pos=options_positions[4]),
               visual.TextStim(win, text='6', height=35, pos=options_positions[5], color=[0, 255, 0]),
               visual.TextStim(win, text='7', height=35, pos=options_positions[6]),
               visual.TextStim(win, text='8', height=35, pos=options_positions[7]),
               visual.TextStim(win, text='9', height=35, pos=options_positions[8]),
               visual.TextStim(win, text='10', height=35, pos=options_positions[9])]

    option7 = [visual.TextStim(win, text='1', height=35, pos=options_positions[0]),
               visual.TextStim(win, text="2", height=35, pos=options_positions[1]),
               visual.TextStim(win, text='3', height=35, pos=options_positions[2]),
               visual.TextStim(win, text='4', height=35, pos=options_positions[3]),
               visual.TextStim(win, text='5', height=35, pos=options_positions[4]),
               visual.TextStim(win, text='6', height=35, pos=options_positions[5]),
               visual.TextStim(win, text='7', height=35, pos=options_positions[6], color=[0, 255, 0]),
               visual.TextStim(win, text='8', height=35, pos=options_positions[7]),
               visual.TextStim(win, text='9', height=35, pos=options_positions[8]),
               visual.TextStim(win, text='10', height=35, pos=options_positions[9])]

    option8 = [visual.TextStim(win, text='1', height=35, pos=options_positions[0]),
               visual.TextStim(win, text="2", height=35, pos=options_positions[1]),
               visual.TextStim(win, text='3', height=35, pos=options_positions[2]),
               visual.TextStim(win, text='4', height=35, pos=options_positions[3]),
               visual.TextStim(win, text='5', height=35, pos=options_positions[4]),
               visual.TextStim(win, text='6', height=35, pos=options_positions[5]),
               visual.TextStim(win, text='7', height=35, pos=options_positions[6]),
               visual.TextStim(win, text='8', height=35, pos=options_positions[7], color=[0, 255, 0]),
               visual.TextStim(win, text='9', height=35, pos=options_positions[8]),
               visual.TextStim(win, text='10', height=35, pos=options_positions[9])]

    option9 = [visual.TextStim(win, text='1', height=35, pos=options_positions[0]),
               visual.TextStim(win, text="2", height=35, pos=options_positions[1]),
               visual.TextStim(win, text='3', height=35, pos=options_positions[2]),
               visual.TextStim(win, text='4', height=35, pos=options_positions[3]),
               visual.TextStim(win, text='5', height=35, pos=options_positions[4]),
               visual.TextStim(win, text='6', height=35, pos=options_positions[5]),
               visual.TextStim(win, text='7', height=35, pos=options_positions[6]),
               visual.TextStim(win, text='8', height=35, pos=options_positions[7]),
               visual.TextStim(win, text='9', height=35, pos=options_positions[8], color=[0, 255, 0]),
               visual.TextStim(win, text='10', height=35, pos=options_positions[9])]

    options_colored = [option0, option1, option2, option3, option4, option5, option6, option7, option8, option9]
    welcome_stimuli = [visual.TextStim(win, text="The experiment is almost over. \n\n You will now be asked a couple of questions about how you experienced the game.\n\n Press SPACE to start.", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text="What do you think the purpose of the experiment was? What did we try to study? \n\n Please input your response. Press ENTER to submit.",height=30, pos=[0, 250], wrapWidth=500),
visual.TextStim(win, text="What were you trying to do when responding to the offers? Did you have any particular goal or strategy?\n\n Press ENTER to submit.",height=30, pos=[0, 250], wrapWidth=500),
visual.TextStim(win, text="How did your strategies differ when you were deciding alone vs. together?\n\n Press ENTER to submit.",height=30, pos=[0, 250], wrapWidth=500),
visual.TextStim(win, text="How would you describe the behavior of the proposer?\n\n Press ENTER to submit.",height=30, pos=[0, 250], wrapWidth=500),
visual.TextStim(win, text="How would you describe the behavior of your partner?\n\n Press ENTER to submit.",height=30, pos=[0, 250], wrapWidth=500),
visual.TextStim(win, text="On a scale from 1-10 how selfish did the proposer seem to you?", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text="How fairly do feel you were treated by the proposer?", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text="How selfish did your partner seem to you?", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text="How fairly do you feel your partner played?", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text="Did you notice anything unusual about the proposer?\n\n Press ENTER to submit.",height=30, pos=[0, 250], wrapWidth=500),
visual.TextStim(win, text="Did you notice anything unusual about your partner?\n\n Press ENTER to submit.",height=30, pos=[0, 250], wrapWidth=500),
visual.TextStim(win, text="How likely do you think it is that the proposer was a computer?", height=30, pos=[0, 0], wrapWidth=500),
visual.TextStim(win, text="The experiment is now over. \n\n Thank you for your participation!", height=30, pos=[0, 0], wrapWidth=500)
]
    stimuli_order = list(range(len(welcome_stimuli)))
    scale_order = [0,1,2,3,4,5,6,7,8,9]
    i = 0
    row_index = 23
    col_index = -1
    welcome_stimuli[stimuli_order[i]].draw()
    win.flip()
    while i < len(stimuli_order)-1:
        if i == 0 or i == len(stimuli_order)-1:
            core.wait(1.0)
            keys = event.getKeys(keyList=['space'])
            if 'space' in keys:
                i += 1
                welcome_stimuli[stimuli_order[i]].draw()
                win.flip()
        if 5 < i < 10 or i == 12:
            welcome_stimuli[stimuli_order[i]].draw()
            not_at_all.draw()
            extremely.draw()
            [i.draw() for i in options]
            win.flip()
            core.wait(1.0)
            keys = event.getKeys(keyList = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9'])
            if any(key in keys for key in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']): #and 5 < i < 10 or i == 12:
                core.wait(1.0)
                sheet.cell(row=row_index, column=1, value=row_index - 1)
                sheet.cell(row=row_index, column=columns[col_index], value=keys[0])
                row_index += 1
                if [int(keys[0]) in scale_order]:
                    scale_num = int(keys[0])
                    if int(keys[0]) > 0:
                        [b.draw() for b in options_colored[scale_order[scale_num]]]
                    if int(keys[0]) == 0:
                        [b.draw() for b in options_colored[scale_order[0]]]
                    welcome_stimuli[stimuli_order[i]].draw()
                    not_at_all.draw()
                    extremely.draw()
                    win.flip()
                    core.wait(1.0)
                    i += 1
                    #print('i after scale', i)
                    if i == 13:
                        break
            else:
                print("invalid key")
        if i not in range(6,10) and i != 12 and i != 0:
            welcome_stimuli[stimuli_order[i]].draw()
            win.flip()
            textbox = visual.TextStim(win, text="", height=20, pos=[0, 0], wrapWidth=500)
            textbox.draw()
            #win.flip()
            while True:
                #win.flip()
                keys = event.getKeys(keyList = ['period', 'comma', 'space', 'return', 'backspace','a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z'])
                if any(key in keys for key in ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', '.', ',']):
                    textbox.text += keys[0]
                if 'return' in keys:
                    print('input:', textbox.text)
                    break
                for key in keys:
                    #if key == 'return':  # Pressing 'return' key ends text input
                    #    break
                    if key == 'backspace':  # Handle backspace
                        #participant_input = participant_input[:-1]
                        #textbox.text = participant_input
                        textbox.text = textbox.text[:-1]
                        textbox.draw()
                        #win.flip()
                        #print('backspace')
                    elif key == 'space':
                        #print(key)
                        space = str(' ')
                        #participant_input += space
                        #textbox.text = participant_input
                        textbox.text += ' '
                        textbox.draw()
                        #win.flip()
                    elif key == 'comma':
                        textbox.text += ','
                        textbox.draw()
                    elif key == 'period':
                        textbox.text += '.'
                        textbox.draw()
                    elif key == 'colon':
                        textbox.text += ':'
                        textbox.draw()
                    else:
                        if len(str(textbox.text)) % 75 == 0:
                            textbox.text += '\n'
                        else:
                            textbox.text = textbox.text[:-1]
                            textbox.text += key
                        #textbox.text = participant_input
                        textbox.draw()
                    welcome_stimuli[stimuli_order[i]].draw()
                    win.flip()

                # Update the text displayed in the Textbox component
                #textbox.text = participant_input
                #textbox.draw()
                #win.flip()

                #if 'return' in keys:
                #    break
            text_responses.append(textbox.text)
            sheet.cell(row=row_index, column=1,value=row_index-1)
            print(text_responses[0])
            sheet.cell(row=row_index, column=columns[col_index], value=text_responses[0])
            row_index += 1
            text_responses.clear()
            i += 1


def save_workbook():
    workbook.save('hexaco_P1.xlsx')
    print('wb saved')

def close_workbook():
    workbook.save('hexaco_P1.xlsx')
    print('wb saved')
    workbook.close()

def thank_you():
    experiment_over = visual.TextStim(win, text="The experiment is now over. \n\n Thank you for your participation!", height=30, pos=[0, 250], wrapWidth=500)
    experiment_over.draw()
    win.flip()
    core.wait(5)

# Close the window at the end
participant_number()
get_empty_col()
welcome()
save_workbook()
bye()
close_workbook()
print("hexaco done")
thank_you()
win.close()
core.quit()